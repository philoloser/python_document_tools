import os
import shutil
import glob
from openpyxl import load_workbook

print(os.getcwd())

workbook = load_workbook('chemiQ.xlsx')

sheet = workbook['BEZP']
cdnkod = []
for row in sheet.iter_rows(min_row=2, max_row=184, min_col=1, max_col=1):  # iteracja przez komórki w kolumnie A
        for cell in row:
            cdnkod.append(cell.value)  # zalaczenie wartosci komorek do listy cdnkod
print(cdnkod)
count = len(cdnkod)
print(count)

sheet = workbook['BEZP']
bielkod = []
for row in sheet.iter_rows(min_row=2, max_row=184, min_col=2, max_col=2):  # iteracja przez komórki w kolumnie A
        for cell in row:
            bielkod.append(cell.value)  # zalaczenie wartosci komorek do listy bielkod
print(bielkod)
count = len(bielkod)
print(count)

source = r'C:\Users\i.janowska\mu_code\_PROJEKTOWE\CHEMIQ\MSDS_R' + '\\'  # r - rawstring, nie rozpoznawanie "\" jako escape char
dest = r'C:\Users\i.janowska\mu_code\_PROJEKTOWE\CHEMIQ\BIELENDA BEZP' + '\\'

output = []
braki = []

for biel_kod, cdn_kod in zip(bielkod, cdnkod):  # iteracja przez obie listy, 1:1 wartości, przez wartości biel_kod
    source_join = os.path.join(source, str(biel_kod))  # określanie ścieżki źródłowej jako [baza] + iterowane {biel_kod} (+zamiana values na string)
    dest_join = os.path.join(dest, str(cdn_kod))  # ścieżka dest, określanie jako baza + iterowanie {cdn_kod}
    destfile = dest_join + '.pdf'

    source_join = source_join + '*'
    source_join = glob.glob(source_join)  # znajodwanie pliku z wildcardem 1234*

    sourcefile = source_join[0]  # wyciąganie pierwszego elementu z listy
    #sourcefile = ''.join(source_join) # zamiana listy glob na string
    #if sourcefile == '':
    braki.append(destfile)

    print('sourcefile: ' + sourcefile)
    print('destfile: ' + destfile)

    output.append(biel_kod)  # appending results of two iterator into a list for counting

    # shutil.copy2(sourcefile, destfile)

count = len(output)
print(count)
print(braki)
