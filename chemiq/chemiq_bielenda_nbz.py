import os
import shutil
import glob
from openpyxl import load_workbook

print(os.getcwd())

workbook = load_workbook('chemiQ.xlsx') #ładowanie excela

sheet = workbook['NIEBEZP']
cdnkod = [] # LISTA KODOW CDN
for row in sheet.iter_rows(min_row=2, max_row=196, min_col=1, max_col=1):  # iteracja przez komórki w kolumnie A
    for cell in row:
        cdnkod.append(cell.value)

brak = [] # LISTA BRAK
tak = []

for row in sheet.iter_rows(min_row=2, max_row=196, min_col=3, max_col=3):  # iteracja przez komórki w kolumnie A
    for cell in row:
        if cell.value == 'TAK':
            tak.append(cell.offset(row=0, column=-2).value) #offsetuj 2 komorki w lewo i odczytaj wartosc
        elif cell.value == 'brak kodu torf':
            brak.append(cell.offset(row=0, column=-2).value)

print(brak)
print(tak)

source = r'C:\Users\i.janowska\mu_code\_PROJEKTOWE\CHEMIQ\BIELENDA NIEBEZP' + '\\'

for tak_v in tak:
    sourcefile = source + tak_v + '.pdf'
    print(sourcefile)
    destfile = source + 'TAK\\' + tak_v + '.pdf'
    print(destfile)

    shutil.move(sourcefile, destfile)

for brak_v in brak:
    sourcefile = source + brak_v + '.pdf'
    print(sourcefile)
    destfile = source + 'BRAK KODU TORF\\' + brak_v + '.pdf'
    print(destfile)

    shutil.move(sourcefile, destfile)