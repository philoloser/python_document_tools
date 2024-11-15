import os
import shutil
import openpyxl
from openpyxl import load_workbook

workbook = load_workbook(
    r'C:\Users\i.janowska\Torf Corporation Sp. z o.o\KJ - _Sharepoint_TC\QC kontrola jakosci\surowce\.Praktyki.staże\2024\CHEMIQ\chemiQ aktualizacja.xlsx')

sheet = workbook['Sheet1']

cdn = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):  # iteracja przez komórki w kolumnie A
    for cell in row:
        cdn.append(cell.value)

print(cdn)

dest = r'C:\Users\i.janowska\Torf Corporation Sp. z o.o\KJ - _Sharepoint_TC\QC kontrola jakosci\surowce\.Praktyki.staże\2024\CHEMIQ\MSDS_CDN\MSDS_CDN_FILTROWANE' + '\\'
source = r'C:\Users\i.janowska\Torf Corporation Sp. z o.o\KJ - _Sharepoint_TC\QC kontrola jakosci\surowce\.Praktyki.staże\2024\CHEMIQ\MSDS_CDN' + "\\"

for plik in cdn:
    sourcefile = source + plik + '.pdf'
    print(sourcefile)

    destfile = dest + plik + '.pdf'
    print(destfile)

    shutil.move(sourcefile, destfile)

dir = []
dir = os.listdir(r'C:\Users\i.janowska\Torf Corporation Sp. z o.o\KJ - _Sharepoint_TC\QC kontrola jakosci\surowce\.Praktyki.staże\2024\CHEMIQ\MSDS_CDN\MSDS_CDN_FILTROWANE')



