import os
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl

cdn = r'C:\Users\i.janowska\mu_code\_PROJEKTOWE\CHEMIQ\MSDS_CDN'
cdn = os.listdir(cdn)
cdnnew = []
for dir in cdn:
    dir = dir.replace('.pdf', '')
    cdnnew.append(dir)

nbzpold = []
nbzp = r'C:\Users\i.janowska\mu_code\_PROJEKTOWE\CHEMIQ\BIELENDA NIEBEZP'

os.walk(nbzp)

# nbzpnew = []
# for dir in nbzp:
#     dir = dir.replace('.pdf', '')
#     nbzpnew.append(dir)

#print(nbzpnew)

#You have to re-allocate the set or re-populate it (re-allocating is easier with a set comprehension):
#new_set = {x.replace('.good', '').replace('.bad', '') for x in set1}



#np.savetxt('LISTACDN.csv', cdnnew, delimiter = ', ', fmt='% s') #csvfile, sourcedata, setdelim, set the format as string



# wb = load_workbook('chemiQ.xlsx')
#
# sheet = Workbook['NIEBEZP']
# nbezp = []
# for row in sheet.iter_rows(min_row=2, maxrow=sheet.max_row, min_col=1, max_col=1):  # iteracja przez komórki w kolumnie A
#     for cell in row:
#         nbezp.append(cell.value)
#
#
# sheet = Workbook['BEZP']
# bezp = []
# for row in sheet.iter_rows(min_row=2, maxrow=sheet.max_row, min_col=1, max_col=1):  # iteracja przez komórki w kolumnie A
#     for cell in row:
#         bezp.append(cell.value)
#
# print(nbezp)